from openpyxl import load_workbook

class TriviaQuestion():
    question: list
    correct_answer: str
    incorrect_answer_1: str
    incorrect_answer_2: str

    def __init__(self, question: list, correct_answer: str, incorrect_answer_1: str, incorrect_answer_2: str):
        self.question = question.copy()
        self.correct_answer = correct_answer
        self.incorrect_answer_1 = incorrect_answer_1
        self.incorrect_answer_2 = incorrect_answer_2

def filter_characters(string: str):
    string = string.lower()
    string = string.replace(" -", "")
    f = "!#$%&/()+-/'"
    for char in f:
        string = string.replace(char, "")
    string = string.replace(" ", "_")
    return string

def process_line(string: str):
    return string.lstrip().rstrip().center(32, " ").rstrip() + "°"
def process_answer(string: str):
    if "°" in string:
        string = string.split("°")
        return f"{string[0].lstrip().rstrip()}°        {string[1].lstrip().rstrip()}°"
    else:
        return string.lstrip().rstrip() + "°°"



wb = load_workbook(filename="DKC2 QUESTIONS.xlsx")

ws = wb.active

category_names = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row+1)]
category_names = list(filter(lambda a: a is not None, category_names))
result_categories = set(category_names)
category_names = [filter_characters(x) for x in result_categories]

dict_categories = {}
for category in result_categories:
    dict_categories[category] = filter_characters(category)

dict_result = {}

for category in set(category_names):
    dict_result[f"trivia_easy_{category}"] = list()
    dict_result[f"trivia_medium_{category}"] = list()
    dict_result[f"trivia_hard_{category}"] = list()

for row in range(2, ws.max_row+1):
    category = filter_characters(ws.cell(row=row, column=2).value)

    result = f'''{'':>4}TriviaQuestion(\n{'':>8}[\n'''
    question = ws.cell(row=row, column=4).value
    question = question.splitlines()
    total_lines = len(question)
    if total_lines == 1:
        question.insert(0, "")
        question.insert(0, "")
        question.append("")
        question.append("")
        question.append("")
    elif total_lines == 2:
        question.insert(0, "")
        question.insert(0, "")
        question.append("")
        question.append("")
    elif total_lines == 3:
        question.insert(0, "")
        question.append("")
        question.append("")
    elif total_lines == 4:
        question.insert(0, "")
        question.append("")
    elif total_lines == 5:
        question.append("")


    for line in question:
        line = process_line(line)
        result += f'''{'':>12}"""{line}""", \n'''
    result += f'''{'':>8}],\n'''
    
    correct_answer = process_answer(str(ws.cell(row=row, column=5).value))
    result += f'''{'':>8}"""{correct_answer}""", \n'''
    incorrect_answer_1 = process_answer(str(ws.cell(row=row, column=6).value))
    result += f'''{'':>8}"""{incorrect_answer_1}""", \n'''
    incorrect_answer_2 = process_answer(str(ws.cell(row=row, column=7).value))
    result += f'''{'':>8}"""{incorrect_answer_2}""", \n'''

    result += f'''{'':>4}),'''

    difficulty = ws.cell(row=row, column=3).value.lower()
    dict_result[f"trivia_{difficulty}_{category}"].append(result)

file_template = '''class TriviaQuestion():
    question: list
    correct_answer: str
    incorrect_answer_1: str
    incorrect_answer_2: str

    def __init__(self, question: list, correct_answer: str, incorrect_answer_1: str, incorrect_answer_2: str):
        self.question = question.copy()
        self.correct_answer = correct_answer
        self.incorrect_answer_1 = incorrect_answer_1
        self.incorrect_answer_2 = incorrect_answer_2

        
trivia_addrs = {
    "easy": [
        0x34F800,
        0x34F900,
        0x34FA00,
        0x34FB00,
        0x34FC00,
        0x34FD00,
    ],
    "medium": [
        0x34F850,
        0x34F950,
        0x34FA50,
        0x34FB50,
        0x34FC50,
        0x34FD50,
    ],
    "hard": [
        0x34F8A0,
        0x34F9A0,
        0x34FAA0,
        0x34FBA0,
        0x34FCA0,
        0x34FDA0,
    ],
}

excluded_questions = [
    2*8, 
    7*8,
    9*8,
    10*8,
    12*8,
    16*8,
    24*8,
    27*8,
    30*8,
    35*8,
    36*8,
    41*8,
    45*8,
]

original_correct_answers = {
    # Galleon
    0: 0,
    1: 0,
    2: 2,
    3: 1,
    4: 1,
    5: 2,
    6: 1,
    7: 0,
    8: 2,
    # Cauldron
    9: 1,
    10: 0,
    11: 0,
    12: 2,
    13: 1,
    14: 0,
    15: 1,
    16: 2,
    17: 2,
    # Quay
    18: 0,
    19: 1,
    20: 2,
    21: 1,
    22: 2,
    23: 1,
    24: 1,
    25: 0,
    26: 2,
    # Kremland
    27: 2,
    28: 0,
    29: 2,
    30: 1,
    31: 1,
    32: 0,
    33: 2,
    34: 1,
    35: 2,
    # Gulch
    36: 0,
    37: 1,
    38: 2,
    39: 2,
    40: 1,
    41: 1,
    42: 2,
    43: 0,
    44: 0,
    # Keep
    45: 1,
    46: 2,
    47: 1,
    48: 0,
    49: 2,
    50: 1,
    51: 1,
    52: 0,
    53: 2,
}
        
'''

with open("Donkey Kong Country 2 - Diddy's Kong Quest (USA) (En,Fr) (Rev 1).sfc", "rb") as f:
    rom = f.read()

rom = bytearray(rom)

original_trivia = {
    "easy": [
    ],
    "medium": [
    ],
    "hard": [
    ],
}

original_correct_answers = {
    # Galleon
    0: 0,
    1: 0,
    2: 2,
    3: 1,
    4: 1,
    5: 2,
    6: 1,
    7: 0,
    8: 2,
    # Cauldron
    9: 1,
    10: 0,
    11: 0,
    12: 2,
    13: 1,
    14: 0,
    15: 1,
    16: 2,
    17: 2,
    # Quay
    18: 0,
    19: 1,
    20: 2,
    21: 1,
    22: 2,
    23: 1,
    24: 1,
    25: 0,
    26: 2,
    # Kremland
    27: 2,
    28: 0,
    29: 2,
    30: 1,
    31: 1,
    32: 0,
    33: 2,
    34: 1,
    35: 2,
    # Gulch
    36: 0,
    37: 1,
    38: 2,
    39: 2,
    40: 1,
    41: 1,
    42: 2,
    43: 0,
    44: 0,
    # Keep
    45: 1,
    46: 2,
    47: 1,
    48: 0,
    49: 2,
    50: 1,
    51: 1,
    52: 0,
    53: 2,
}

excluded_questions = [
    2*8, 
    7*8,
    9*8,
    10*8,
    12*8,
    16*8,
    24*8,
    27*8,
    30*8,
    35*8,
    36*8,
    41*8,
    45*8,
]

start_addr = 0x34C591

for idx in range(0, 0x1B0, 8):
    if idx in excluded_questions:
        continue
    offset = int.from_bytes(rom[start_addr+idx:start_addr+idx+2], "little")
    offset |= 0x370000
    text_data = rom[offset:offset+0xD0]     # The max amount should be around 0xC6
    text_data = text_data.split(b'\x00')
    question = []
    for i in range(6):
        question.append(text_data.pop(0).decode() + "°")

    offset = int.from_bytes(rom[start_addr+idx+4:start_addr+idx+6], "little")
    offset |= 0x370000
    text_data = rom[offset:offset+0x70]     # The max amount should be around 0x66
    text_data = text_data.split(b'\x00')
    answers = []
    for i in range(3):
        text = text_data.pop(0).decode() + "°" + text_data.pop(0).decode() + "°"
        text = text[8:]
        answers.append(text)

    idy = original_correct_answers[idx >> 3]
    correct_answer = answers.pop(idy)
    data = TriviaQuestion(question, correct_answer, answers.pop(0), answers.pop(0))
    if idx % 0x48 < 0x10:
        original_trivia["easy"].append(data)
    elif idx % 0x48 < 0x28:
        original_trivia["medium"].append(data)
    else:
        original_trivia["hard"].append(data)
    

with open("A:/Archipelago/worlds/dkc2/data/Trivia.py", "w", encoding="utf-8") as f:
    r = file_template

    sorted_dict = dict(sorted(dict_result.items()))

    for category, questions in sorted_dict.items():
        r += f'''{category} = [\n'''
        for question in questions:
            r += f"{question}\n"
        r += f"]\n\n"

    r += f'''\ntrivia_data = {{\n'''
    for key, value in dict_categories.items():
        r += f'''{'':>4}"{key}": [\n{'':>8}trivia_easy_{value}, \n{'':>8}trivia_medium_{value}, \n{'':>8}trivia_hard_{value},\n{'':>4}],\n'''
    r += '}\n\n'

    f.write(r)

